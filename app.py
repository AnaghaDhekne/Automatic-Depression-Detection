import os
from werkzeug.utils import secure_filename
from flask import Flask, render_template, redirect, url_for, session, request, Response, flash
from flask_bootstrap import Bootstrap
from flask_wtf import FlaskForm
from wtforms import StringField, PasswordField, IntegerField, RadioField
from wtforms.validators import InputRequired, Email, Length, NumberRange
from werkzeug.security import generate_password_hash, check_password_hash
import mysql.connector
import cv2
import time
from datetime import date, datetime
from password_strength import PasswordPolicy
import shutil
import random
from flask_mail import Mail, Message
from openpyxl import Workbook
import openpyxl
import threading
from itsdangerous import URLSafeTimedSerializer
import face_recognition
from keras.models import model_from_json
import numpy as np
import matplotlib.pyplot as plt
from fpdf import FPDF
from email.mime.multipart import MIMEMultipart
import smtplib, ssl
from email import encoders
from email.mime.base import MIMEBase
from email.mime.text import MIMEText

app = Flask(__name__)
UPLOAD_FOLDER = 'C:/Users/anagh/OneDrive/Pictures/Uploads'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
APP_ROOT = os.path.dirname(os.path.abspath(__file__))
app.config['SECRET_KEY'] = 'Thisissupposedtobesecret!'
app.config['IS_LOGIN'] = False
bootstrap = Bootstrap(app)
mail_settings = {
    "MAIL_SERVER": 'smtp.gmail.com',
    "MAIL_PORT": 465,
    "MAIL_USE_TLS": False,
    "MAIL_USE_SSL": True,
    "MAIL_USERNAME": 'depressiede@gmail.com',
    "MAIL_PASSWORD": 'RAAS@123'
}
global RUN
RUN = 0
global EMAIL
global transaction_id
CONNECTION = mysql.connector.connect(
    host="localhost",
    user="root",
    password="root",
    database="project",
    auth_plugin='mysql_native_password'
)
app.config['CONNECTION'] = CONNECTION
app.config['MAIL_DEFAULT_SENDER'] = 'depressiede@gmail.com'
app.config['SECURITY_PASSWORD_SALT'] = 'my_precious_two'
app.config.update(mail_settings)
mail = Mail(app)
s = URLSafeTimedSerializer('Thisissupposedtobesecret!')


class LoginForm(FlaskForm):
    email = StringField('Email', validators=[InputRequired(), Email(message='Invalid email'), Length(max=50)])
    password = PasswordField('Password', validators=[InputRequired(), Length(min=8, max=80)])


class OTPForm(FlaskForm):
    otp = IntegerField('OTP', validators=[InputRequired(), NumberRange(00000, 99999, 'Wrong OTP')])


class EmailForm(FlaskForm):
    email = StringField('Email', validators=[InputRequired(), Email(message='Invalid email'), Length(max=50)])


class ChangePasswordForm(FlaskForm):
    password = PasswordField('Password', validators=[InputRequired(), Length(min=8, max=80)])


class RegisterForm(FlaskForm):
    email = StringField('Email', validators=[InputRequired(), Email(message='Invalid email'), Length(max=50)])
    username = StringField('Username', validators=[InputRequired(), Length(min=4, max=15)])
    password = PasswordField('Password', validators=[InputRequired(), Length(min=8, max=80)])
    age = IntegerField('Age', validators=[InputRequired(), NumberRange(18, 75, 'Age not supported for application')])
    gender = RadioField('Gender', choices=[(1, 'Male'), (2, 'Female')], coerce=int)


class FieldsRequiredForm(FlaskForm):
    """Require all fields to have content. This works around the bug that WTForms radio
    fields don't honor the `DataRequired` or `InputRequired` validators.
    """

    class Meta:
        def render_field(self, field, render_kw):
            render_kw.setdefault('required', True)
            return super().render_field(field, render_kw)


class QuestionForm(FieldsRequiredForm):
    question1 = RadioField(
        '1. Sadness', choices=[(1, 'I do not feel sad.'),
                               (2, 'I feel sad much of the time.'),
                               (3, 'I am sad all the time.'),
                               (4, 'I am so sad or unhappy that I cannot stand it.')], coerce=int,
        validators=[InputRequired()])

    question2 = RadioField(
        '2. Pessimism', choices=[(1, 'I am not discouraged about my future.'),
                                 (2, 'I feel more discouraged about my future than I used to.'),
                                 (3, 'I do not expect things to work out for me.'),
                                 (4, 'I feel my future is hopeless and will only get worse.')], coerce=int,
        validators=[InputRequired()])

    question3 = RadioField(
        '3. Past Failure', choices=[(1, 'I do not feel like a failure.'),
                                    (2, 'I have failed more than I should have.'),
                                    (3, 'As I look back, I see a lot of failures.'),
                                    (4, 'I feel I am a total failure as a person.')], coerce=int,
        validators=[InputRequired()])

    question4 = RadioField(
        '4. Loss of Pleasure', choices=[(1, 'I get as much pleasure as I ever did from the things I enjoy.'),
                                        (2, 'I do not enjoy things as much as I used to.'),
                                        (3, 'I get very little pleasure from the things I used to enjoy.'),
                                        (4, 'I cannot get any pleasure from the things I used to enjoy.')], coerce=int,
        validators=[InputRequired()])

    question5 = RadioField(
        '5. Guilty Feelings', choices=[(1, 'I do not feel particularly guilty.'),
                                       (2, 'I feel guilty over many things I have done or should have done.'),
                                       (3, 'I feel quite guilty most of the time.'),
                                       (4, 'I feel guilty all of the time.')], coerce=int, validators=[InputRequired()])

    question6 = RadioField(
        '6. Punishment Feelings', choices=[(1, 'I do not feel I am being punished.'),
                                           (2, 'I feel I may be punished.'),
                                           (3, 'I expect to be punished.'),
                                           (4, 'I feel I am being punished.')], coerce=int,
        validators=[InputRequired()])

    question7 = RadioField(
        '7. Self-Dislike', choices=[(1, 'I feel the same about myself as ever.'),
                                    (2, 'I have lost confidence in myself.'),
                                    (3, 'I am disappointed in myself.'),
                                    (4, 'I dislike myself.')], coerce=int, validators=[InputRequired()])

    question8 = RadioField(
        '8. Self-Criticalness', choices=[(1, 'I do not criticize or blame myself more than usual.'),
                                         (2, 'I am more critical of myself than I used to be.'),
                                         (3, 'I criticize myself for all of my faults.'),
                                         (4, 'I blame myself for everything bad that happens.')], coerce=int,
        validators=[InputRequired()])

    question9 = RadioField(
        '9. Suicidal Thoughts or Wishes', choices=[(1, 'I do not have any thoughts of killing myself.'),
                                                   (2,
                                                    'I have thoughts of killing myself, but I would not carry them out.'),
                                                   (3, 'I would like to kill myself.'),
                                                   (4, 'I would kill myself if I had the chance.')], coerce=int,
        validators=[InputRequired()])

    question10 = RadioField(
        '10. Crying', choices=[(1, 'I do not cry anymore than I used to.'),
                               (2, 'I cry more than I used to.'),
                               (3, 'I cry over every little thing.'),
                               (4, 'I feel like crying, but I cannot.')], coerce=int, validators=[InputRequired()])

    question11 = RadioField(
        '11. Agitation', choices=[(1, 'I am no more restless or wound up than usual.'),
                                  (2, 'I feel more restless or wound up than usual.'),
                                  (3, 'I am so restless or agitated, it is hard to stay still.'),
                                  (4, 'I am so restless or agitated that I have to keep moving or doing something.')],
        coerce=int, validators=[InputRequired()])
    question12 = RadioField(
        '12. Loss of Interest', choices=[(1, 'I have not lost interest in other people or activities.'),
                                         (2, 'I am less interested in other people or things than before.'),
                                         (3, 'I have lost most of my interest in other people or things.'),
                                         (4, 'It is hard to get interested in anything.')], coerce=int,
        validators=[InputRequired()])

    question13 = RadioField(
        '13. Indecisiveness', choices=[(1, 'I make decisions about as well as ever.'),
                                       (2, 'I find it more difficult to make decisions than usual.'),
                                       (3, 'I have much greater difficulty in making decisions than I used to.'),
                                       (4, 'I have trouble making any decisions.')], coerce=int,
        validators=[InputRequired()])

    question14 = RadioField(
        '14. Worthlessness', choices=[(1, 'I do not feel I am worthless.'),
                                      (2, 'I do not consider myself as worthwhile and useful as I used to.'),
                                      (3, 'I feel more worthless as compared to others.'),
                                      (4, 'I feel utterly worthless.')], coerce=int, validators=[InputRequired()])
    question15 = RadioField(
        '15. Loss of Energy', choices=[(1, 'I have as much energy as ever.'),
                                       (2, 'I have less energy than I used to have.'),
                                       (3, 'I do not have enough energy to do very much.'),
                                       (4, 'I do not have enough energy to do anything.')], coerce=int,
        validators=[InputRequired()])
    question16 = RadioField(
        '16. Changes in Sleeping Pattern', choices=[(1, 'I can sleep as well as usual. '),
                                                    (2, 'I do not sleep as well as I used to.'),
                                                    (3,
                                                     'I wake up 1-2 hours earlier than usual and find it hard to get back to sleep.'),
                                                    (4,
                                                     'I wake up several hours earlier than I used to and cannot get back to sleep.')],
        coerce=int, validators=[InputRequired()])
    question17 = RadioField(
        '17. Irritability', choices=[(1, 'I am not more irritable than usual.'),
                                     (2, 'I am more irritable than usual.'),
                                     (3, 'I am much more irritable than usual.'),
                                     (4, 'I am irritable all the time.')], coerce=int, validators=[InputRequired()])
    question18 = RadioField(
        '18. Changes in Appetite', choices=[(1, 'My appetite is no worse than usual.'),
                                            (2, 'My appetite is not as good as it used to be.'),
                                            (3, 'My appetite is much worse now.'),
                                            (4, 'I have no appetite at all anymore.')], coerce=int,
        validators=[InputRequired()])
    question19 = RadioField(
        '19. Concentration Difficulty', choices=[(1, 'I can concentrate as well as ever.'),
                                                 (2, 'I cannot concentrate as well as usual.'),
                                                 (3, 'It is hard to keep my mind on anything for very long.'),
                                                 (4, 'I find I cannot concentrate on anything.')], coerce=int,
        validators=[InputRequired()])
    question20 = RadioField(
        '20. Tiredness or Fatigue', choices=[(1, 'I am no more tired or fatigued than usual.'),
                                             (2, 'I get more tired or fatigued more easily than usual.'),
                                             (3, 'I am too tired or fatigued to do a lot of the things I used to do.'),
                                             (4, 'I am too tired or fatigued to do most of the things I used to do.')],
        coerce=int, validators=[InputRequired()])
    question21 = RadioField(
        '21. Loss of Interest in Sex', choices=[(1, 'I have not noticed any recent change in my interest in sex.'),
                                                (2, 'I am less interested in sex than I used to be.'),
                                                (3, 'I am much less interested in sex now.'),
                                                (4, 'I have lost interest in sex completely.')], coerce=int,
        validators=[InputRequired()])


@app.errorhandler(404)
def not_found(e):
    return render_template("404.html")


@app.errorhandler(500)
def handle_intsrverr(e):
    return render_template("500.html")


@app.errorhandler(408)
def handle_reqTimeOut(e):
    return render_template("408.html")


@app.errorhandler(503)
def handle_serverUnavailable(e):
    return render_template("503.html")


@app.route('/pagenotfunderror', methods=['GET', 'POST'])
def pagenotfunderror():
    return redirect(url_for('index'))


@app.route('/signup', methods=['GET', 'POST'])
def signup():
    form = RegisterForm()
    login_form = LoginForm()
    policy = PasswordPolicy.from_names(
        length=8,  # min length: 8
        uppercase=1,  # need min. 2 uppercase letters
        numbers=1,  # need min. 2 digits
        special=1,  # need min. 2 special characters
    )
    if not policy.test(form.password.data):
        if form.validate_on_submit():
            error = 0
            mydb = CONNECTION
            my_cursor = mydb.cursor()
            sql = "SELECT email FROM users"
            my_cursor.execute(sql)
            email_user = my_cursor.fetchall()
            for email_id in email_user:
                if email_id[0] == form.email.data:
                    error = 2
                    break
            if error == 2:
                flash('The mail id is already been used. Please use other email', 'warning')
                return render_template('Signuppage.html', form=form)
            hashed_password = generate_password_hash(form.password.data, method='sha256')
            if form.gender.data == 1:
                gender = 'Male'
            else:
                gender = 'Female'
            mycursor = mydb.cursor()
            sql = "INSERT INTO users (username,email,password,age,gender,confirmed) VALUES (%s,%s,%s,%s,%s,%s)"
            val = (form.username.data, form.email.data, hashed_password, form.age.data, gender, 0)
            mycursor.execute(sql, val)

            token = generate_confirmation_token(form.email.data)
            confirm_url = url_for('confirm_email', token=token, _external=True)
            html = render_template('activate.html', confirm_url=confirm_url)
            subject = "Please confirm your email"
            send_email(form.email.data, subject, html)
            flash('A confirmation email has been sent via email.', 'success')
            mydb.commit()
            return redirect(url_for('index'))
    flash('Enter the information as following rules: strong password, age more than 18', 'warning')
    return render_template('Signuppage.html', form=form)


# part of Email Verification module
def generate_confirmation_token(email):
    serializer = URLSafeTimedSerializer(app.config['SECRET_KEY'])
    return serializer.dumps(email, salt=app.config['SECURITY_PASSWORD_SALT'])


# part of Email Verification module
def send_email(to, subject, template):
    msg = Message(
        subject,
        recipients=[to],
        html=template,
        sender=app.config['MAIL_DEFAULT_SENDER']
    )
    mail.send(msg)


# part of Email Verification module
@app.route('/confirm/<token>')
def confirm_email(token):
    form = RegisterForm()
    mydb = CONNECTION
    mycursor = mydb.cursor()
    session['session_email'] = form.email.data
    try:
        email = confirm_token(token)
    except:
        flash('The confirmation link is invalid or has expired.', 'danger')
        return redirect(url_for('index'))
    finally:
        sql = "UPDATE users SET confirmed = %s WHERE email = %s"
        val = (1, email)
        mycursor.execute(sql, val)
        mydb.commit()
        flash('Account confirmed')
    return redirect(url_for('index'))


# part of Email Verification module
def confirm_token(token, expiration=3600):
    serializer = URLSafeTimedSerializer(app.config['SECRET_KEY'])
    try:
        email = serializer.loads(
            token,
            salt=app.config['SECURITY_PASSWORD_SALT'],
            max_age=expiration
        )
    except:
        return False
    return email


@app.route('/signuppage', methods=['GET', 'POST'])
def signuppage():
    form = RegisterForm()
    return render_template('Signuppage.html', form=form)


@app.route('/login', methods=['GET', 'POST'])
def login():
    form = LoginForm()
    if form.validate_on_submit():
        mydb = CONNECTION
        mycursor = mydb.cursor()
        sql = "SELECT * FROM users WHERE email = %s"
        email = form.email.data
        mycursor.execute(sql, (email,))
        user = mycursor.fetchone()
        now = datetime.now()
        current_time = now.strftime("%H:%M:%S")
        if user:
            if user[6] == 1:
                if check_password_hash(user[3], form.password.data):
                    app.config['IS_LOGIN'] = True
                    session['loggedin'] = True
                    session['emailId'] = user[2]
                    sql = "INSERT INTO login_info (email,password,date,time) VALUES (%s,%s,%s,%s)"
                    val = (form.email.data, user[3], date.today(), current_time)
                    mycursor.execute(sql, val)
                    mydb.commit()
                    select_query = "SELECT id FROM login_info ORDER BY id DESC LIMIT 1"
                    mycursor.execute(select_query)
                    id_array = mycursor.fetchone()
                    session['id'] = id_array[0]
                    global EMAIL
                    EMAIL = session['emailId']
                    delete_previous_folder()
                    delete_face_database()
                    return redirect(url_for('captureimage'))
                else:
                    flash(
                        "Weak Password or Incorrect Password! Must be combination of letters, numbers, symbols. Try again",
                        "warning")
                    return render_template('index.html', lform=form)
            else:
                token = generate_confirmation_token(email)
                confirm_url = url_for('confirm_email', token=token, _external=True)
                html = render_template('activate.html', confirm_url=confirm_url)
                subject = "Please confirm your email"
                send_email(form.email.data, subject, html)
                flash('Mail has not been confirmed. A confirmation email has been sent via email.', 'warning')
                return render_template('index.html', lform=form)
        flash('Not a valid user. Please Sign Up', 'warning')
        return render_template('index.html', lform=form)
    flash("Enter correct mail id and password", "warning")
    return redirect(url_for('index'))


# part of forget password module
@app.route('/passwordsubmit', methods=['GET', 'POST'])
def password_submit():
    login_form = LoginForm()
    cform = ChangePasswordForm()
    if cform.validate_on_submit():
        mydb = CONNECTION
        mycursor = mydb.cursor()
        hashed_password = generate_password_hash(cform.password.data, method='sha256')
        print(hashed_password)
        sql = "UPDATE users SET password = %s WHERE email = %s"
        email = session['password_update_mail']
        print(email)
        val = (hashed_password, email)
        mycursor.execute(sql, val)
        mydb.commit()
        with app.app_context():
            msg = Message(subject="Password Updated",
                          sender=app.config.get("MAIL_USERNAME"),
                          recipients=[email],  # replace with your email for testing
                          body="Your Password has been updated. If you have not requested a change of password then send a mail to depressiede@gmail.com")
            mail.send(msg)
        session.pop('password_update_mail', None)
    flash("Password Updated", "success")
    return render_template('index.html', lform=login_form)


# part of forget password module
@app.route('/otpsubmit', methods=['GET', 'POST'])
def otp_submit():
    cform = ChangePasswordForm()
    oform = OTPForm()
    eform = EmailForm()
    login_form = LoginForm()
    if oform.validate_on_submit() and (session['otp'] == oform.otp.data):
        session.pop('otp', None)
        return render_template('forgetpassword.html', oform=oform, eform=eform, lform=login_form,
                               value=2, cform=cform)
    flash("Wrong one time password", "warning")
    return render_template('index.html', lform=login_form)


# part of forget password module
@app.route('/emailsubmit', methods=['GET', 'POST'])
def email_submit():
    oform = OTPForm()
    eform = EmailForm()
    login_form = LoginForm()
    cform = ChangePasswordForm()
    if eform.validate_on_submit():
        mydb = CONNECTION
        mycursor = mydb.cursor()
        sql = "SELECT * FROM users WHERE email = %s"
        email = eform.email.data
        mycursor.execute(sql, (email,))
        user = mycursor.fetchone()
        try:
            if user[2] == eform.email.data:
                with app.app_context():
                    otp = random.randrange(00000, 99999)
                    msg = Message(subject="Password OTP",
                                  sender=app.config.get("MAIL_USERNAME"),
                                  recipients=[eform.email.data],  # replace with your email for testing
                                  body="Your One Time Password is: " + str(otp))
                    mail.send(msg)
                    session['otp'] = otp
                    session['password_update_mail'] = eform.email.data
                    return render_template('forgetpassword.html', oform=oform, eform=eform, lform=login_form,
                                           value=1, cform=cform)
        except:
            flash("Email id not registered else check internet connection", "warning")
            return render_template('index.html', lform=login_form)
    flash("Enter correct mail id : '....@gmail.com'", "warning")
    return render_template('index.html', lform=login_form)


@app.route('/forgetpassword', methods=['GET', 'POST'])
def forgetpassword():
    oform = OTPForm()
    eform = EmailForm()
    login_form = LoginForm()
    cform = ChangePasswordForm()
    return render_template('forgetpassword.html', oform=oform, eform=eform, lform=login_form,
                           value=0, cform=cform)


# called in video_stream function
def get_frame():
    camera_port = 0
    camera = cv2.VideoCapture(camera_port)  # this makes a web cam object
    ret, img = camera.read()
    imgencode = cv2.imencode('.jpg', img)[1]
    stringData = imgencode.tostring()
    yield (b'--frame\r\n'
           b'Content-Type: text/plain\r\n\r\n' + stringData + b'\r\n')
    cv2.imwrite('face.jpg', img)
    del camera


@app.route('/video_stream')
def video_stream():
    return Response(get_frame(), mimetype='multipart/x-mixed-replace; boundary=frame')


# to capture real time image
@app.route('/captureimage', methods=['GET', 'POST'])
def captureimage():
    return render_template('start.html', value=0)


# to submit real time image
@app.route('/submit', methods=['GET', 'POST'])
def submit():
    # return render_template('start.html', value=1)
    return render_template('start.html', value=1)


@app.route('/uploadvideo', methods=['GET', 'POST'])
def uploadvideo():
    return render_template('UploadVideo.html')


# to delete the folder of uploads when old user logs in again, called in login function
def delete_previous_folder():
    global EMAIL
    directory = EMAIL
    path = os.path.join(UPLOAD_FOLDER, directory)
    if os.path.isdir(path):
        shutil.rmtree(path, ignore_errors=True)
        time.sleep(5)


def delete_face_database():
    current_path = os.getcwd()
    path = os.path.join(current_path, 'face_database')
    if os.path.isdir(path):
        for filename in os.listdir(path):
            os.remove(os.path.join(path, filename))
        time.sleep(5)


@app.route("/upload", methods=["POST", "GET"])
def upload():
    mydb = CONNECTION
    solve_form = QuestionForm()
    if 'file' not in request.files:
        flash('No File Part', 'warning')
        return render_template('UploadVideo.html')
    uploaded_files = request.files.getlist('file')
    file = request.files['file']
    if file.filename == '':
        flash('No Selected File', 'warning')
        return render_template('UploadVideo.html')
    directory = session['emailId']
    path = os.path.join(UPLOAD_FOLDER, directory)
    os.mkdir(path)
    image_directory = os.path.join(path, 'Image')
    os.mkdir(image_directory)
    user_image = os.path.join(image_directory, 'user_face.jpg')
    cv2.imwrite(user_image, cv2.imread('face.jpg'))
    os.remove('face.jpg')
    video_directory = os.path.join(path, 'Video')
    os.mkdir(video_directory)
    xl_directory = os.path.join(path, 'Xl')
    os.mkdir(xl_directory)
    for file in uploaded_files:
        filename = secure_filename(file.filename)
        file.save(os.path.join(video_directory, filename))
    now = datetime.now()
    current_time = now.strftime("%H:%M:%S")
    mycursor = mydb.cursor()
    sql = "INSERT INTO video_table (video_id,email,date,time,number_of_files,status,folder) VALUES (%s,%s,%s,%s,%s,%s,%s)"
    val = (session['id'], session['emailId'], date.today(), current_time, len(uploaded_files), 0, path)
    mycursor.execute(sql, val)
    mydb.commit()
    mycursor1 = mydb.cursor()
    sql1 = "SELECT transactionid FROM video_table ORDER BY video_id DESC LIMIT 1"
    mycursor1.execute(sql1)
    id_array = mycursor1.fetchone()
    global transaction_id
    transaction_id = id_array[0]
    return render_template('SolveQuestionaire.html', form=solve_form)


@app.route('/solve_questionaire', methods=['post', 'get'])
def submit_action():
    form = QuestionForm()
    if form.validate_on_submit():
        score = return_score(form.question1.data) + return_score(form.question2.data) + return_score(
            form.question3.data) + return_score(form.question4.data) + return_score(form.question5.data) + return_score(
            form.question6.data) + return_score(form.question7.data) + return_score(form.question8.data) + return_score(
            form.question9.data) + return_score(form.question10.data) + return_score(
            form.question11.data) + return_score(form.question12.data) + return_score(
            form.question13.data) + return_score(form.question14.data) + return_score(
            form.question15.data) + return_score(form.question16.data) + return_score(
            form.question17.data) + return_score(form.question18.data) + return_score(
            form.question19.data) + return_score(form.question20.data) + return_score(form.question21.data)
        now = datetime.now()
        current_time = now.strftime("%H:%M:%S")
        mydb = CONNECTION
        mycursor = mydb.cursor()
        sql = "INSERT INTO bdi_table (id,email,date,time,score) VALUES (%s,%s,%s,%s,%s)"
        val = (session['id'], session['emailId'], date.today(), current_time, score)
        mycursor.execute(sql, val)
        mydb.commit()
        workbook = Workbook()
        sheet = workbook.active
        sheet.cell(row=1, column=1).value = return_score(form.question1.data)
        sheet.cell(row=2, column=1).value = return_score(form.question2.data)
        sheet.cell(row=3, column=1).value = return_score(form.question3.data)
        sheet.cell(row=4, column=1).value = return_score(form.question4.data)
        sheet.cell(row=5, column=1).value = return_score(form.question5.data)
        sheet.cell(row=6, column=1).value = return_score(form.question6.data)
        sheet.cell(row=7, column=1).value = return_score(form.question7.data)
        sheet.cell(row=8, column=1).value = return_score(form.question8.data)
        sheet.cell(row=9, column=1).value = return_score(form.question9.data)
        sheet.cell(row=10, column=1).value = return_score(form.question10.data)
        sheet.cell(row=11, column=1).value = return_score(form.question11.data)
        sheet.cell(row=12, column=1).value = return_score(form.question12.data)
        sheet.cell(row=13, column=1).value = return_score(form.question13.data)
        sheet.cell(row=14, column=1).value = return_score(form.question14.data)
        sheet.cell(row=15, column=1).value = return_score(form.question15.data)
        sheet.cell(row=16, column=1).value = return_score(form.question16.data)
        sheet.cell(row=17, column=1).value = return_score(form.question17.data)
        sheet.cell(row=18, column=1).value = return_score(form.question18.data)
        sheet.cell(row=19, column=1).value = return_score(form.question19.data)
        sheet.cell(row=20, column=1).value = return_score(form.question20.data)
        sheet.cell(row=21, column=1).value = return_score(form.question21.data)
        sheet.cell(row=22, column=1).value = score
        directory = session['emailId']
        path = os.path.join(UPLOAD_FOLDER, directory)
        xl_directory = os.path.join(path, 'Xl')
        xl_file = os.path.join(xl_directory, 'bdi.xlsx')
        workbook.save(xl_file)
    else:
        print(form.errors)
    return render_template('employeeresult.html', form=form)


# to return bdi score of the radio button selected
def return_score(data):
    if data == 1:
        return 0
    elif data == 2:
        return 1
    elif data == 3:
        return 2
    else:
        return 3


# called in depression detection function to get face_database
def load_images_from_folder(folder):
    images = []
    for filename in os.listdir(folder):
        img = cv2.imread(os.path.join(folder, filename))
        if img is not None:
            images.append(img)
    return images


def generate_report(transaction_id, EMAIL):
    my_db = CONNECTION
    my_cursor = my_db.cursor()
    sql = "SELECT * FROM results WHERE transaction_id = %s "
    my_cursor.execute(sql, (transaction_id,))
    data = my_cursor.fetchone()
    my_db.commit()
    score = data[2]
    negative = data[4]
    positive = data[6]
    neutral = data[5]
    level = data[8]
    rmail = data[9]

    my_db1 = CONNECTION
    my_cursor1 = my_db1.cursor()
    sql1 = "SELECT * FROM users WHERE email = %s"
    my_cursor1.execute(sql1, (EMAIL,))
    data1 = my_cursor1.fetchone()
    my_db1.commit()
    name = data1[1]
    age = data1[4]
    gender = data1[5]

    labels = ['Positive', 'Negative', 'Neutral']
    sizes = [positive, negative, neutral]  # Add upto 100%
    # Plot the pie chart
    plt.pie(sizes, labels=labels, autopct='%1.1f%%', startangle=90)
    # Equal aspect ratio ensures that pie is drawn as a circle.
    plt.axis('equal')
    # Display the graph onto the screen
    plt.savefig('abc.png')

    class PDF(FPDF):
        def header(self):
            self.set_font('Arial', 'B', 15)
            # create a cell
            self.cell(200, 10, txt="De Depressie", ln=1, align='C')
            # add another cell
            self.cell(200, 10, txt="Depression Detection Result", ln=2, align='C')
            self.ln(5)

        def footer(self):
            # Position at 1.5 cm from bottom
            self.set_y(-15)
            # Arial italic 8
            self.set_font('Arial', 'I', 8)
            # Text color in gray
            self.set_text_color(128)
            # Page number
            self.cell(0, 10, 'Page ' + str(self.page_no()), 0, 0, 'C')

        def topic_title(self, label):
            # Arial 12
            self.set_font('Arial', '', 12)
            # Background color
            self.set_fill_color(200, 220, 255)
            # Title
            self.cell(0, 6, '%s  ' % (label), 0, 1, 'L', 1)
            # Line break
            self.ln(5)

        def topic_body(self, name):
            # Read text file
            with open(name, 'rb') as fh:
                txt = fh.read().decode('latin-1')
            # Times 12
            self.set_font('Arial', '', 11)
            # Output justified text
            self.multi_cell(0, 5, txt)
            # Line break
            self.ln()

        def result_body(self):
            self.set_font('Arial', '', 11)
            self.cell(0, 6, 'Name : %s' % (name), 0, 1)
            self.cell(0, 6, 'Age : %d' % (age), 0, 1)
            self.cell(0, 6, 'Gender : %s' % (gender), 0, 1)

            self.cell(0, 6,
                      'Based on the visual inputs and your BDI score which is %d you have %s Level of depression. LEVEL 1- MINIMAL, LEVEL 2- MILD, LEVEL 3- MODERATE, LEVEL 4- SEVERE ' % (
                          score, level))
            self.ln()

            if level == 'Minimal':
                tname = 'minimal.txt'
            elif level == 'Mild':
                tname = 'mild.txt'
            elif level == 'Moderate':
                tname = 'moderate.txt'
            else:
                tname = 'severe.txt'

            # Read text file
            with open(tname, 'rb') as fh:
                txt = fh.read().decode('latin-1')
            # Times 12
            self.set_font('Arial', '', 11)
            # Output justified text
            self.multi_cell(0, 5, txt)
            # Line break
            self.ln()

        def print_chapter(self):
            self.add_page()
            self.topic_title('Who are we?')
            self.topic_body('Note.txt')

            self.topic_title('Result :')
            self.result_body()

            self.topic_title('Emotions Detected :')
            if level == 'Minimal':
                self.image('abc.png', None, None, 200, 118);
            elif level == 'Mild':
                self.image('abc.png', None, None, 200, 110);
            elif level == 'Moderate':
                self.image('abc.png', None, None, 180, 95);
            else:
                self.image('abc.png', None, None, 180, 85);

            self.add_page()
            self.topic_title('FAQ :')
            self.topic_body('FAQ.txt')

            self.topic_title('Contact Details :')
            self.set_font('Arial', '', 11)
            self.cell(0, 6, 'Feel free to contact us on the depressiedegmail.com mail ', 0, 1)

    pdf = PDF()
    pdf.set_author('Team De Depressie')
    pdf.print_chapter()
    pdf.output('result.pdf', 'F')

    subject = "Depression Level Detection Report"
    body = "Hello %s,\nYour test report is ready.Thank you for your time. \nPlease find the attachment sent from De Depressie Team.\n \nWarm Regards,\nTeam De Depressie." % (
        name)
    sender_email = 'depressiede@gmail.com'
    password = 'RAAS@123'

    # Create a multipart message and set headers
    message = MIMEMultipart()
    message["From"] = sender_email
    message["To"] = rmail
    message["Subject"] = subject
    message["Bcc"] = rmail  # Recommended for mass emails

    # Add body to email
    message.attach(MIMEText(body, "plain"))
    filename = "result.pdf"  # In same directory as script

    # Open PDF file in binary mode
    with open(filename, "rb") as attachment:
        # Add file as application/octet-stream
        # Email client can usually download this automatically as attachment
        part = MIMEBase("application", "octet-stream")
        part.set_payload(attachment.read())

    # Encode file in ASCII characters to send by email
    encoders.encode_base64(part)

    # Add header as key/value pair to attachment part
    part.add_header(
        "Content-Disposition",
        f"attachment; filename= {filename}",
    )

    # Add attachment to message and convert message to string
    message.attach(part)
    text = message.as_string()
    print(sender_email, password)
    # Log in to server using secure context and send email
    context = ssl.create_default_context()
    with smtplib.SMTP_SSL("smtp.gmail.com", 465, context=context) as server:
        server.login(sender_email, password)
        server.sendmail(sender_email, rmail, text)
    delete_graphs()


def path_file_delete(path):
    if os.path.isfile(path):
        os.remove(path)
        time.sleep(5)


def delete_graphs():
    current_path = os.getcwd()
    path = os.path.join(current_path, 'result.pdf')
    path_file_delete(path)
    path = os.path.join(current_path, 'abc.png')
    path_file_delete(path)


def correlation_function(emotions, bdi):
    negative_group = 0
    positive_group = 0
    neutral_group = 0
    bdi_score = bdi[21]
    for frame in emotions:
        if emotions[frame] == 'Angry' or emotions[frame] == 'Disgust' or emotions[frame] == 'Fear' or emotions[
            frame] == 'Sad':
            negative_group = negative_group + 1
        if emotions[frame] == 'Surprise' or emotions[frame] == 'Happy':
            positive_group = positive_group + 1
        if emotions[frame] == 'Neutral':
            neutral_group = neutral_group + 1
    total = negative_group + positive_group + neutral_group
    bdi_percentage = ((bdi_score / 63) * 100)
    negative_percentage = ((negative_group / total) * 100)
    positive_percentage = ((positive_group / total) * 100)
    neutral_percentage = ((neutral_group / total) * 100)
    level_percentage = (bdi_percentage + negative_percentage) / 2
    level = 0
    if level_percentage <= 25:
        level = 1
    elif level_percentage > 25 or level_percentage <= 50:
        level = 2
    elif level_percentage > 50 or level_percentage <= 75:
        level = 3
    else:
        level = 4
    print(total, bdi_score, bdi_percentage)
    print(negative_group, negative_percentage)
    print(positive_group, positive_percentage)
    print(neutral_group, neutral_percentage)
    print(level_percentage, level)

    mydb3 = CONNECTION
    mycursor3 = mydb3.cursor(buffered=True)
    global transaction_id
    sql = "UPDATE video_table SET status = %s WHERE transactionid = %s"
    val = (1, transaction_id)
    mycursor3.execute(sql, val)
    mydb3.commit()

    global EMAIL
    mydb5 = CONNECTION
    mycursor5 = mydb5.cursor()
    sql_select = "SELECT * from users WHERE email = %s"
    mycursor5.execute(sql_select, (EMAIL,))
    user = mycursor5.fetchone()
    mydb5.commit()

    mydb4 = CONNECTION
    mycursor4 = mydb4.cursor()
    sql_insert = "INSERT INTO results (transaction_id,bdi_score,bdi_percentage,negative_percentage,neutral_percentage,positive_percentage,result_percentage,level,email_id,age,gender) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"
    val = (transaction_id, bdi_score, bdi_percentage, negative_percentage, neutral_percentage, positive_percentage,
           level_percentage, level, EMAIL, user[4], user[5])
    mycursor4.execute(sql_insert, val)
    mydb4.commit()
    generate_report(transaction_id, EMAIL)


def depression_detection():
    # get path of folder from database
    global EMAIL
    global transaction_id
    mydb = CONNECTION
    mycursor = mydb.cursor()
    sql = "SELECT * FROM video_table WHERE transactionid = %s"
    mycursor.execute(sql, (transaction_id,))
    user = mycursor.fetchone()
    path = user[7]

    # if path exists
    if not os.path.isdir(path):
        print("directory doesnot exist")
        with app.app_context():
            msg = Message(subject="De Depressie Results!!",
                          sender=app.config.get("MAIL_USERNAME"),
                          recipients=[EMAIL],  # replace with your email for testing
                          body="""Hi,
                          The De Depressie team is sorry to inform you that your results are not generated as the complete process of submission was not completed in your last attempt. To get your results do remember to complete steps as follows:
1. Capture Image
2. Upload Video
3. Solve Questionnaire
If any of the step is not followed then you do not get any results. For any queries feel free to mail us at depressiede@gmail.com
                          
Thank you,
Team De Depressie.""")
            mail.send(msg)
    else:
        # check if bdi xl exists
        xl_directory = os.path.join(path, 'Xl')
        if len(os.listdir(xl_directory)) == 0:
            print('no xl file')
            with app.app_context():
                msg = Message(subject="De Depressie Results!!",
                              sender=app.config.get("MAIL_USERNAME"),
                              recipients=[EMAIL],  # replace with your email for testing
                              body="""Hi,
                                      The De Depressie team is sorry to inform you that your results are not generated as the complete process of submission was not completed in your last attempt. To get your results do remember to complete steps as follows:
    1. Capture Image
    2. Upload Video
    3. Solve Questionnaire
    If any of the step is not followed then you do not get any results. For any queries feel free to mail us at depressiede@gmail.com

    Thank you,
    Team De Depressie.""")
                mail.send(msg)
        else:
            xl_file = os.path.join(xl_directory, 'bdi.xlsx')
            wb_obj = openpyxl.load_workbook(xl_file)
            sheet_obj = wb_obj.active
            bdi = []
            for i in range(1, 23):
                cell_obj = sheet_obj.cell(row=i, column=1).value
                bdi.append(cell_obj)

            image_directory = os.path.join(path, 'Image')
            for filename in os.listdir(image_directory):
                name = filename
            image = os.path.join(image_directory, name)
            lmm_image = face_recognition.load_image_file(image)
            lmm_face_encoding = face_recognition.face_encodings(lmm_image)[0]
            known_faces = [
                lmm_face_encoding
            ]

            current_path = os.getcwd()
            counter = 0
            video_directory = os.path.join(path, 'Video')
            for filename in os.listdir(video_directory):
                frame_number = 0
                input_movie = cv2.VideoCapture(os.path.join(video_directory, filename))
                length = int(input_movie.get(cv2.CAP_PROP_FRAME_COUNT))
                while True:
                    # Grab a single frame of video
                    ret, frame = input_movie.read()
                    frame_number += 1
                    # Quit when the input video file ends
                    if not ret:
                        break
                    # Find all the faces and face encodings in the current frame of video
                    face_locations = face_recognition.face_locations(frame)
                    face_encodings = face_recognition.face_encodings(frame, face_locations)
                    face_names = []
                    for face_encoding in face_encodings:
                        # See if the face is a match for the known face(s)
                        match = face_recognition.compare_faces(known_faces, face_encoding, tolerance=0.50)
                        name = None
                        if match[0]:
                            name = "user"
                        face_names.append(name)
                    # Label the results
                    for (top, right, bottom, left), name in zip(face_locations, face_names):
                        if not name:
                            continue
                        # Draw a box around the face
                        cv2.rectangle(frame, (left, top), (right, bottom), (0, 0, 255), 2)
                        crop_img = frame[top:bottom, left:right]
                        # image_enhancer.load_images_from_folder()
                        if "user" == name:
                            # normalframes
                            cv2.imwrite(current_path + "/face_database/" + "user" + str(counter) + ".jpg", crop_img)
                            counter = counter + 1
                    print("Writing frame {} / {}".format(frame_number, length))
                input_movie.release()
                cv2.destroyAllWindows()

            # get images from face_database to input to cnn
            time.sleep(15)
            current_path = os.getcwd()
            image_input = load_images_from_folder(os.path.join(current_path, 'face_database'))
            print("Image Loaded")

            if len(image_input) < 21:
                print('not enough input')
                with app.app_context():
                    msg = Message(subject="De Depressie Results!!",
                                  sender=app.config.get("MAIL_USERNAME"),
                                  recipients=[EMAIL],  # replace with your email for testing
                                  body="""Hi,
                                          The De Depressie team is sorry to inform you that your results are not generated as the videos provided where not enough to reach a result. To get your results do remember to complete steps as follows:
        1. Capture Image
        2. Upload Video
        3. Solve Questionnaire
        If any of the step is not followed then you do not get any results. For any queries feel free to mail us at depressiede@gmail.com

        Thank you,
        Team De Depressie.""")
                    mail.send(msg)
            else:
                # load the machine learning model
                json_file = open('fer.json', 'r')
                loaded_model_json = json_file.read()
                json_file.close()
                loaded_model = model_from_json(loaded_model_json)
                loaded_model.load_weights("fer.h5")

                # add label array
                labels = ['Angry', 'Disgust', 'Fear', 'Happy', 'Sad', 'Surprise', 'Neutral']
                emotions = {}
                count = 0

                for full_size_image in image_input:
                    gray = cv2.cvtColor(full_size_image, cv2.COLOR_RGB2GRAY)
                    cropped_img = np.expand_dims(np.expand_dims(cv2.resize(gray, (48, 48)), -1), 0)
                    cv2.normalize(cropped_img, cropped_img, alpha=0, beta=1, norm_type=cv2.NORM_L2, dtype=cv2.CV_32F)
                    yhat = loaded_model.predict(cropped_img)
                    label = labels[int(np.argmax(yhat))]
                    emotions[count] = label
                    count = count + 1
                print(emotions)
                correlation_function(emotions, bdi)
                # correlation logic db insertion of result

        delete_previous_folder()
        delete_face_database()


# generation of thread to run second part of project
@app.before_first_request
def activate_job():
    def run_job():
        global RUN
        while True:
            print(RUN)
            if 1 == RUN:
                depression_detection()
                RUN = 0
                break
            if RUN == 0:
                print("Run recurring task")
                time.sleep(60)

    thread = threading.Thread(target=run_job)
    thread.start()


@app.route('/logout')
def logout():
    global RUN
    RUN = 1
    global EMAIL
    EMAIL = session['emailId']
    app.config['IS_LOGIN'] = False
    session.pop('loggedin', None)
    session.pop('id', None)
    session.pop('emailId', None)
    return redirect(url_for('index'))


@app.route('/')
def index():
    login_form = LoginForm()
    if app.config['IS_LOGIN']:
        return redirect(url_for('uploadvideo'))
    else:
        return render_template('index.html', lform=login_form, error=0)


if __name__ == '__main__':
    app.run(debug=False)
